import pyupbit
import time
import openpyxl
import os

#과거데이터 조회
count = 4000
timeunit = 'minute5' #1,3,5,10,30,60m day week month
timesec = 5*60
os.mkdir(timeunit)

tickers = pyupbit.get_tickers(fiat="KRW")
number = len(tickers)
done = 0
for ticker in tickers:
    try:
        data = pyupbit.get_ohlcv(ticker=ticker, interval=timeunit, count=count, to=None, period=0.1)
        now = time.time() - count * timesec
        data_list = data.columns.values.tolist() + data.values.tolist()
        wb = openpyxl.Workbook()
        datasheet = wb.worksheets[0]
        for i in range(0, 6):
            datasheet.cell(row=1, column=i + 2, value=data_list[i])
        for i in range(6, 6 + count):
            for j in range(0, 6):
                datasheet.cell(row=i - 4, column=1,
                               value=time.strftime('%Y-%m-%d %H:%M', time.localtime(now + timesec * (i - 5))))
                datasheet.cell(row=i - 4, column=j + 2, value=data_list[i][j])
        wb.save(timeunit + '/' + ticker + '.xlsx')
        done += 1
        print(str(done) + '/' + str(number) + ' 완료')
    except:
        done += 1
        print(ticker + ' 저장실패')
        print(str(done) + '/' + str(number) + ' 완료')