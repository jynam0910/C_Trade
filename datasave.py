import pyupbit
import time
import openpyxl

#TODO 200개 이상 저장하도록 구현
#TODO 종목 여러개 한번에 하도록 구현

#과거데이터 조회
count = 200
data = pyupbit.get_ohlcv(ticker="KRW-BTC", interval="day", count=count)
#엑셀정리
now = time.time() - count*24*60*60
data_list = data.columns.values.tolist() + data.values.tolist()
wb = openpyxl.Workbook()
datasheet = wb.worksheets[0]
for i in range(0,6):
    datasheet.cell(row=1, column=i+2, value=data_list[i])
for i in range(6,6 + count):
    for j in range(0,6):
        datasheet.cell(row=i-4, column=1, value=time.strftime('%Y-%m-%d %I:%M:%S', time.localtime(now + 24*60*60*(i-5))))
        datasheet.cell(row=i-4, column=j+2, value=data_list[i][j])
wb.save('test.xlsx')

'''k=30
sum = 0
for i in range(6,6+k):
    sum = sum + data_list[i][3]
avg = sum/k

status = 'buy'
money = 0
price = 0
for i in range(6+k,206):
    if data_list[i][3] >= avg and data_list[i-1][3] < avg and status == 'buy':
        price = data_list[i][3]
        status = 'sell'
        print(str(price) + "에 구매함")
    elif data_list[i][3] < data_list[i-1][3] and status == 'sell':
        money = money + data_list[i][3] - price
        status = 'buy'
        print(str(data_list[i][3]) + "에 판매함")
    else:
        print('status = ' + status + " 현재가: " + str(data_list[i][3]), 'avg: ' + str(avg))
    sum = 0
    for i in range(i+1-k,i+1):
        sum = sum + data_list[i][3]
    avg = sum / k
print(money)'''